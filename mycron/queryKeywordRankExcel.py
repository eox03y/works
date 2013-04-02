import sys
import time
import datetime
from operator import itemgetter
import codecs
import re
import handolUtil
import handolS3Util

def line_yield(mem_lines):
	for line in mem_lines.splitlines():
		line = unicode(line, encoding='utf-8')
		yield line

def get_keyvalue_from_logline(fields):
	d = {}
	for fld in fields:
		try:
			key,value = fld.split('=', 1)
		except:
			return None
		value=value.strip()
		if value=='': value="Nil"
		d[key]=value
	return d

MAPPER_SHOPID_APPS = {
	"450":"KOR", "234":"GBR", "235":"GBR"
}
def filer_for_SA_query(line):
	''' input: line of SamsungApps query log
		output: list of fields
	'''
	flds = line.split('|@')
	d = get_keyvalue_from_logline(flds[1:])
	if d==None: return None
	qh = d.get('qh', None)
	q = d.get('q', None)
	shop = d.get('shp', None)
	if qh==None or q==None or shop==None: return None
	
	try:
		ctry = MAPPER_SHOPID_APPS[shop]		
	except:
		return None
	tot = d.get('tot', None)
	if tot==None: tot = 0
	else: tot = int(tot)	
	return [qh, q, ctry, tot]

def get_shopid_num(strval):
	pos = strval.find('SHOPID:(("')
	if pos < 0: return ''
	shopstr = strval[10:]

shop_re = re.compile(r'"(\d+)"', re.UNICODE)
def get_shopid_num(strval):
	match = shop_re.search(strval)
	if match:
		return match.group(1)
	return ''


MAPPER_SHOPID_LHRHVH = {
	"GLH" : { "1":"KOR", "3":"GBR" },
	"GRH" : { "24":"KOR", "26":"GBR" },
	"GVH" : { "1007":"KOR", "1006":"GBR" }
}

def filer_for_LHRHVH_query(line):
	''' input: line of LH/RH/VH query log
		output: list of fields
		shop id:
			GLH: KOR-1, GBR-3
			GRH: KOR-24, GBR-26
			GVH: KOR-1007, GBR-1006
		log line) fld 9 - query, fld 11 - shop id, fld 18 - SVC(LH/LH)
		06-03-2013 00:21:28:598|null|@null|@null|@null|@null|@null|@null|@null|@historia |@b0|@SHOPID:(("32")):;channelId:(("9")):;deviceId:(("5"))|@null|@10|@0|@|@|@|@GRH|@252323317|@31|@title_it_auto:(historia*)^1.0|@47
	'''
	flds = line.split('|@')
	if len(flds) < 22: return None

	q = flds[8].strip()
	shopfld = flds[10]	# SHOPID:(("32"))

	svc = flds[17][:3] # GRH, GLH, GVH_COMMON
	qh = flds[18] # query hash code
	solrq = flds[20] # title_it_auto:(historia*)^1.0
	tot = flds[21]

	if solrq[9:13]=='auto': return None

	shopid = get_shopid_num(shopfld)
	if q==None or svc==None: return None
	try:
		ctry = MAPPER_SHOPID_LHRHVH[svc][shopid]
	except:
		return None

	if tot==None: tot = 0
	else: tot = int(tot)

	return [qh, q, ctry, tot, svc]


###
class LogAnalyzer:
	def __init__(self, svcname, myname):
		'''
		svcname must be one of "LHRHVH", "APPS".
		'''
		self.SVC = svcname
		self.myname = myname
		self.errcnt_q = 0
		self.errcnt_c = 0
		self.srcList = []
		self.fileList = {}

	def querylog_from_s3(self, bucket, s3file):
		org_size, mem = handolS3Util.get_file_to_mem( handolS3Util.conn, bucket, s3file)
		print "Open S3: %s/%s" % (bucket, s3file)
		print "ORG [%d] LOAD [%d]" % (org_size, len(mem))
		if org_size==0 or org_size != len(mem):
			print "Failed:", s3file
			return

		res = self.querylog_from_memory(mem)
		self.fileList[s3file] = res

	def querylog_from_memory(self, mem):
		res = self.iter_query_lines(line_yield(mem))
		return res

	def querylog_from_file(self, fname):
		try:
			fp = codecs.open(fname, 'rb', encoding='utf-8')
			print "Open:", fname
		except:
			print "Open Failed:", fname
			return

		res = self.iter_query_lines(fp)
		fp.close()
		self.fileList[fname] = res

	def iter_query_lines(self, lineiter):
		if self.SVC=="APPS":
			filter_func = filer_for_SA_query
		elif self.SVC=="LHRHVH":
			filter_func = filer_for_LHRHVH_query
		else:
			return 0

		"""
		for line in lineiter:
			flds = filter_func(line)
			if flds==[]:
				continue
			self.srcList.append(flds)
		"""
		# this is more 'Pythonic'
		# oneList - result from one file
		oneList = [flds for flds in [filter_func(line) for line in lineiter] if flds]
		self.srcList += oneList
		return len(oneList)

	#
	def get_stats(self):
		""" calculate statistics and store into self.statD
		"""
		self.statD = {}
		for v in self.srcList:
			keyword = v[1].strip().lower()
			ctry = v[2]
			tot = v[3]
			if tot==0: nores = 1
			else: nores = 0
			if self.SVC=="APPS": svc = "APPS"
			else: svc = v[4]

			SVC_CTRY = "%s-%s" % (svc, ctry)
			kdict = self.statD.get(SVC_CTRY, None) # dict by Keywords
			if not kdict:
				kdict = handolUtil.AddValueDict(SVC_CTRY)
				self.statD[SVC_CTRY] = kdict

			kdict.add(keyword, [1, tot, nores])

	#
	def print_info(self, outf=''):
		"""
		print out result.
		if outf is given, stdout is redirected to the file 'outf'
		"""
		if outf != '':
			backup = sys.stdout
			sys.stdout = open(outf, 'wb')

		for SVC_CTRY,kdict in self.statD.iteritems():
			print "SVC_CTRY:,", SVC_CTRY
			#print "searches(top 1500) , ", kdict.sum_of_tops(1500)
			N = 1500
			print "# Top %s" % (N)
			kdict.prn(N)

		if outf != '':	sys.stdout = backup

	#
	def	write_info(self):
		for SVC_CTRY,kdict in self.statD.iteritems():
			fname = "%s.%s.%s.dat" % (self.SVC, self.myname, SVC_CTRY)
			print "Save to:", fname
			N = 1500
			kdict.save_key_last(fname, N)
	#
	def	write_excel(self):
		try:
			import openpyxl
		except:
			print "cannot import openpyxl"
			return
		wb = openpyxl.workbook.Workbook()
		ws = wb.get_active_sheet()
		fname = "%s.%s.xlsx" % (self.SVC, self.myname)
		print "Save to:", fname
		N = 1500
		cnt_sheet = 0
		for SVC_CTRY,kdict in self.statD.iteritems():
			# Sheet
			ws.title = SVC_CTRY
			kdict.rank()
			row = 0
			for key, val in kdict.ranked:
				# one row
				ws.cell(row=row, column=0).value = val[0]
				ws.cell(row=row, column=1).value = val[1]/val[0]
				ws.cell(row=row, column=2).value = val[2]
				ws.cell(row=row, column=3).value = key
				row += 1
				if row >= N: break
			cnt_sheet += 1
			if cnt_sheet < len(self.statD):
				ws = wb.create_sheet()

		wb.save(filename=fname)

		# put excel file to S3
		time.sleep(0.2)
		handolS3Util.put_file( handolS3Util.conn, 
			'sch-emr', fname, 'statReport/%s' % fname)
				
				
##

def test_w_local_LHRHVH():
	analyzer = LogAnalyzer("LHRHVH", "any")
	stopwatch = handolUtil.StopWatch()

	analyzer.querylog_from_file("osp_query.log.20130306")
	print "Loading & ETL: %f sec" % (stopwatch.laptime())

	analyzer.get_stats()
	print "Calc Stats: %f sec" % (stopwatch.laptime())
	analyzer.write_info()

def test_w_local_APPS():
	analyzer = LogAnalyzer("APPS", "any")
	stopwatch = handolUtil.StopWatch()

	analyzer.querylog_from_file("euospsch03.2.osp_query.log.20130215")
	#analyzer.querylog_from_file("a.query.log")
	print "Loading & ETL: %f sec" % (stopwatch.laptime())

	analyzer.get_stats()
	print "Calc Stats: %f sec" % (stopwatch.laptime())
	analyzer.write_info()


######
def do_many_days_s3_LHRHVH(argv):
	if len(argv) < 3:
		print "usage: start_day end_day"
		print "usage: 2011-10-7 2011-10-13"
		sys.exit()

	daylist = handolUtil.get_day_list(argv[1],argv[2])
	analyzer = LogAnalyzer("LHRHVH", "%s-%s" %(daylist[0], daylist[-1]))
	stopwatch = handolUtil.StopWatch()
	for daystr in daylist:
		s3_file = "DEVELOPING/app/7nmc1m75ij/hubs1-log/query_log/osp_query.log.%s" % (daystr)
		analyzer.querylog_from_s3("sch-emr", s3_file)
		print "Loading & ETL: %f sec" % (stopwatch.laptime())
		s3_file = "DEVELOPING/app/7nmc1m75ij/hubs2-log/query_log/osp_query.log.%s" % (daystr)
		analyzer.querylog_from_s3("sch-emr", s3_file)
		print "Loading & ETL: %f sec" % (stopwatch.laptime())

	analyzer.get_stats()
	print "Calc Stats: %f sec" % (stopwatch.laptime())
	#analyzer.write_info()
	analyzer.write_excel()

######
def do_many_days_s3_APPS(argv):
	if len(argv) < 3:
		print "usage: start_day end_day"
		print "usage: 2011-10-7 2011-10-13"
		sys.exit()

	QUERY_FILES = [
		'euospcomp07.osp_query.log',
		'euospcomp08.osp_query.log',
		'euospsch01.osp_query.log',
		'euospsch03.osp_query.log',
		'euospsch01.2.osp_query.log',
		'euospsch03.2.osp_query.log',
	]
	CLICK_FILES = [
		'euospsch03.osp_click.log',
		'euospcomp08.osp_click.log'
	]

	#QUERY_FILES = QUERY_FILES[:2]

	daylist = handolUtil.get_day_list(argv[1],argv[2])
	analyzer = LogAnalyzer("APPS", "%s-%s" %(daylist[0], daylist[-1]))
	stopwatch = handolUtil.StopWatch()
	for daystr in daylist:
		for q in QUERY_FILES:
			s3_file = "DEVELOPING/app/7nmc1m75ij/apps-log/query_log/%s/%s.%s" % (daystr, q, daystr)
			analyzer.querylog_from_s3("sch-emr", s3_file)
			print "Loading & ETL: %f sec" % (stopwatch.laptime())

	analyzer.get_stats()
	print "Calc Stats: %f sec" % (stopwatch.laptime())
	#analyzer.write_info()
	analyzer.write_excel()



if __name__=="__main__":
	#test_w_local_LHRHVH()
	#test_w_local_APPS()
	if len(sys.argv) < 4:
		print "usage: SVC_NAME start_day end_day"
		print "usage: APPS 2011-10-7 2011-10-13"
		print "usage: LHRHVH 2011-10-7 2011-10-13"
		sys.exit()
	if sys.argv[1]=='APPS':
		do_many_days_s3_APPS(sys.argv[1:])
	else:
		do_many_days_s3_LHRHVH(sys.argv[1:])
